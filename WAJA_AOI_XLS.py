import re
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

# --- CONFIGURATION ---
INPUT_FILE = "lotwreport.adi"
BANDS_LIST = ["160M", "80M", "40M", "30M", "20M", "17M", "15M", "12M", "10M", "6M"]

PREFECTURES = {
    "01": "Hokkaido", "02": "Aomori", "03": "Iwate", "04": "Akita", "05": "Yamagata",
    "06": "Miyagi", "07": "Fukushima", "08": "Niigata", "09": "Nagano", "10": "Tokyo",
    "11": "Kanagawa", "12": "Chiba", "13": "Saitama", "14": "Ibaraki", "15": "Tochigi",
    "16": "Gunma", "17": "Yamanashi", "18": "Shizuoka", "19": "Gifu", "20": "Aichi",
    "21": "Mie", "22": "Kyoto", "23": "Shiga", "24": "Nara", "25": "Osaka",
    "26": "Wakayama", "27": "Hyogo", "28": "Toyama", "29": "Fukui", "30": "Ishikawa",
    "31": "Okayama", "32": "Shimane", "33": "Yamaguchi", "34": "Tottori", "35": "Hiroshima",
    "36": "Kagawa", "37": "Tokushima", "38": "Ehime", "39": "Kochi", "40": "Fukuoka",
    "41": "Saga", "42": "Nagasaki", "43": "Kumamoto", "44": "Oita", "45": "Miyazaki",
    "46": "Kagoshima", "47": "Okinawa"
}

def stream_adif_records(file_path):
    current_record = []
    try:
        with open(file_path, "r", encoding="latin-1") as f:
            for line in f:
                current_record.append(line)
                if "<eor>" in line.lower():
                    yield "".join(current_record)
                    current_record = []
    except FileNotFoundError:
        print(f"[-] Error: File '{file_path}' not found!")
        return

def parse_adif_field(record, field_name):
    pattern = rf"<{field_name}:?\d*>[^<\s]+"
    match = re.search(pattern, record, re.IGNORECASE)
    if match: return match.group(0).split(">")[-1].strip()
    return None

def extract_prefecture(record):
    for field in ["STATE", "CNTY"]:
        val = parse_adif_field(record, field)
        if val:
            match = re.match(r"^(\d{1,2})", val)
            if match: return f"{int(match.group(1)):02d}"
    return None

def apply_excel_formatting(file_path, sheet_name):
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = writer.book
            worksheet = workbook[sheet_name]
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006", bold=True)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            green_font = Font(color="006100")
            
            # Formatiranje kolone "Status" (C)
            for cell in worksheet["C"][1:]:
                if cell.value == "MISSING":
                    cell.fill = red_fill
                    cell.font = red_font
                elif cell.value == "WORKED":
                    cell.fill = green_fill
                    cell.font = green_font
    except Exception as e:
        print(f"[!] Warning: Could not style Excel file: {e}")

def main():
    matrix_data = {code: {b: None for b in BANDS_LIST} for code in PREFECTURES.keys()}
    
    for record in stream_adif_records(INPUT_FILE):
        if parse_adif_field(record, "DXCC") == "339":
            band = (parse_adif_field(record, "BAND") or "").upper()
            pref_code = extract_prefecture(record)
            callsign = parse_adif_field(record, "CALL") or ""
            
            if band in BANDS_LIST and pref_code in PREFECTURES:
                is_portable = "/P" in callsign.upper()
                display_call = f"*{callsign}" if is_portable else callsign
                
                # Upisujemo u matricu ako još nemamo potvrdu za taj band
                if matrix_data[pref_code][band] is None:
                    qso_date = parse_adif_field(record, "QSO_DATE") or ""
                    time_on = parse_adif_field(record, "TIME_ON") or ""
                    
                    matrix_data[pref_code][band] = {
                        "Callsign": display_call,
                        "Mode": (parse_adif_field(record, "MODE") or "UNK").upper(),
                        "Date": f"{qso_date[:4]}-{qso_date[4:6]}-{qso_date[6:8]}" if len(qso_date)==8 else qso_date,
                        "Time": f"{time_on[:2]}:{time_on[2:4]}" if len(time_on)>=4 else time_on,
                        "IsPortable": is_portable
                    }

    # 1. POJEDINAČNI FAJLOVI PO BANDOVIMA
    for band in BANDS_LIST:
        band_rows = []
        for code, pref_name in sorted(PREFECTURES.items()):
            data = matrix_data[code][band]
            band_rows.append({
                "Code": code, "Prefecture": pref_name, "Status": "WORKED" if data else "MISSING",
                "Callsign": data["Callsign"] if data else "", "Mode": data["Mode"] if data else "", 
                "Date": data["Date"] if data else "", "Time": data["Time"] if data else ""
            })
        df_band = pd.DataFrame(band_rows)
        fn = f"{band.lower()}_report.xlsx"
        df_band.to_excel(fn, index=False, sheet_name=f"{band} Report")
        apply_excel_formatting(fn, f"{band} Report")

    # 2. OBJEDINJENI FAJL
    all_rows = [{"Band": b, "Code": c, "Prefecture": p, "Status": "WORKED" if matrix_data[c][b] else "MISSING",
                 "Callsign": matrix_data[c][b]["Callsign"] if matrix_data[c][b] else "",
                 "Mode": matrix_data[c][b]["Mode"] if matrix_data[c][b] else "",
                 "Date": matrix_data[c][b]["Date"] if matrix_data[c][b] else "",
                 "Time": matrix_data[c][b]["Time"] if matrix_data[c][b] else ""}
                for b in BANDS_LIST for c, p in sorted(PREFECTURES.items())]
    pd.DataFrame(all_rows).to_excel("all_JA_per_bands.xlsx", index=False, sheet_name="All Bands")
    apply_excel_formatting("all_JA_per_bands.xlsx", "All Bands")

    # 3. PORTABLE /P LISTA
    portable_rows = []
    for b in BANDS_LIST:
        for c, p in sorted(PREFECTURES.items()):
            d = matrix_data[c][b]
            if d and d.get("IsPortable"):
                portable_rows.append({"Band": b, "Code": c, "Prefecture": p, "Callsign": d["Callsign"], 
                                      "Mode": d["Mode"], "Date": d["Date"], "Time": d["Time"], "Note": ""})
    
    if portable_rows:
        df_port = pd.DataFrame(portable_rows)
        df_port.to_excel("portable_stations_report.xlsx", index=False, sheet_name="Portable")
        print("[+] Kreirani izveštaji. /P lista za provjeru: portable_stations_report.xlsx")
    else:
        print("[*] Nema /P stanica za listu.")

# --- 4. KREIRANJE MATRIČNOG IZVEŠTAJA (WAJA_Summary.xlsx) ---
    print("[*] Kreiranje matričnog izvještaja (WAJA_Summary.xlsx)...")
    
    matrix_rows = []
    for code, pref_name in sorted(PREFECTURES.items()):
        row = {"Code": code, "Prefecture": pref_name}
        for band in BANDS_LIST:
            data = matrix_data[code][band]
            # Upisujemo samo Callsign ako postoji, inače "."
            row[band] = data["Callsign"] if data else "."
        matrix_rows.append(row)
    
    df_matrix = pd.DataFrame(matrix_rows)
    df_matrix.to_excel("WAJA_Summary.xlsx", index=False)
    
    # Dodatno formatiranje za ovaj fajl (bojenje ćelija sa ".")
    try:
        with pd.ExcelWriter("WAJA_Summary.xlsx", engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = writer.book
            worksheet = workbook["Sheet1"] # Default ime
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Prolazimo kroz sve ćelije od kolone C do kraja
            for col in range(3, len(BANDS_LIST) + 3):
                for row in range(2, len(PREFECTURES) + 2):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value == ".":
                        cell.fill = red_fill
        print("[+] Kreiran matrični izvještaj: WAJA_Summary.xlsx")
    except Exception as e:
        print(f"[!] Warning: Could not style WAJA_Summary.xlsx: {e}")
        
if __name__ == "__main__":
    main()